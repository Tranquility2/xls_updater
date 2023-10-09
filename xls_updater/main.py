"""Module providing util to convert xls to newer xlsx."""
import click
import xlrd
from openpyxl.workbook import Workbook


def change_file_extension_with_string_methods(filename, new_extension):
    """Function updating the given filename string to a given extention."""
    if "." in filename:
        name, old_extension = filename.rsplit(".", 1)
        print(f"Replacing old '{old_extension}' with new '{new_extension}' extension")
        new_filename = name + "." + new_extension
    else:
        new_filename = filename + "." + new_extension
    return new_filename


@click.command()
@click.argument("src_file_path", type=click.Path(exists=True))
def cvt_xls_to_xlsx(src_file_path):
    """Function converting the given xls file to the newer xlsx format."""
    dst_file_path = change_file_extension_with_string_methods(src_file_path, "xlsx")
    print(f"Output={dst_file_path}")
    book_xls = xlrd.open_workbook(src_file_path)
    book_xlsx = Workbook()

    sheet_names = book_xls.sheet_names()
    for sheet_index, sheet_name in enumerate(sheet_names):
        sheet_xls = book_xls.sheet_by_name(sheet_name)
        if sheet_index == 0:
            sheet_xlsx = book_xlsx.active
            sheet_xlsx.title = sheet_name
        else:
            sheet_xlsx = book_xlsx.create_sheet(title=sheet_name)

        for row in range(0, sheet_xls.nrows):
            for col in range(0, sheet_xls.ncols):
                sheet_xlsx.cell(row=row + 1, column=col + 1).value = sheet_xls.cell_value(row, col)

    book_xlsx.save(dst_file_path)


if __name__ == "__main__":
    cvt_xls_to_xlsx()  # pylint: disable=no-value-for-parameter
