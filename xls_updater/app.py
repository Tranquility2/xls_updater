"""Module to convert xls to newer xlsx."""

import pathlib

import xlrd
from openpyxl.workbook import Workbook
from tqdm import tqdm


def convert_xls_to_xlsx(src_file_path: pathlib.Path) -> None:
    """Function converting the given xls file to the newer xlsx format."""
    dst_file_path = src_file_path.with_suffix(".xlsx")
    book_xls = xlrd.open_workbook(src_file_path)
    book_xlsx = Workbook()

    sheet_names = book_xls.sheet_names()

    total_cells = 0
    for sheet_name in sheet_names:
        sheet_xls = book_xls.sheet_by_name(sheet_name)
        total_cells += sheet_xls.nrows * sheet_xls.ncols

    with tqdm(total=total_cells, desc="Converting xls to xlsx") as progress_bar:
        for sheet_index, sheet_name in enumerate(sheet_names):
            sheet_xls = book_xls.sheet_by_name(sheet_name)
            if sheet_index == 0:
                sheet_xlsx = book_xlsx.active
                sheet_xlsx.title = sheet_name
            else:
                sheet_xlsx = book_xlsx.create_sheet(title=sheet_name)

            for row in range(0, sheet_xls.nrows):
                for col in range(0, sheet_xls.ncols):
                    sheet_xlsx.cell(row=row + 1, column=col + 1).value = sheet_xls.cell_value(row, col)
                    progress_bar.update(1)

    book_xlsx.save(dst_file_path)
