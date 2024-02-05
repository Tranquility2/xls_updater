"""Test app."""

from pathlib import Path

import openpyxl
import xlwt

from xls_updater.app import convert_xls_to_xlsx


def create_test_excel_file() -> None:
    """Create test excel file."""
    book = xlwt.Workbook()
    sheet = book.add_sheet("Sheet")
    sheet.write(0, 0, 42)
    sheet = book.add_sheet("Sheet2")
    sheet.write(0, 0, "test")
    book.save("sample.xls")


def test_convert_xls_to_xlsx() -> None:
    """Test convert_xls_to_xlsx."""
    create_test_excel_file()
    convert_xls_to_xlsx(Path("sample.xls"))
    assert Path("sample.xlsx").exists()
    # check content
    book_xlsx = openpyxl.load_workbook("sample.xlsx")
    assert book_xlsx.sheetnames == ["Sheet", "Sheet2"]
    assert book_xlsx["Sheet"]["A1"].value == 42
    assert book_xlsx["Sheet2"]["A1"].value == "test"
    Path("sample.xls").unlink()
    Path("sample.xlsx").unlink()
