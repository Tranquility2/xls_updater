"""Test app."""
from pathlib import Path

import openpyxl
import xlwt
from click.testing import CliRunner

from xls_updater.cli import cli


def create_test_excel_file() -> None:
    """Create test excel file."""
    book = xlwt.Workbook()
    sheet = book.add_sheet("Sheet")
    sheet.write(0, 0, 42)
    sheet = book.add_sheet("Sheet2")
    sheet.write(0, 0, "test")
    book.save("sample.xls")


def test_cli() -> None:
    """Test app."""
    runner = CliRunner()

    with runner.isolated_filesystem():
        # generate sample excel file
        create_test_excel_file()
        # convert it to xls
        result = runner.invoke(cli, ["sample.xls"])
        assert result.exit_code == 0, result.output
        assert result.output == "Replacing old 'xls' with new 'xlsx' extension\nOutput=sample.xlsx\n"
        assert Path("sample.xlsx").exists()
        # check content
        book_xlsx = openpyxl.load_workbook("sample.xlsx")
        assert book_xlsx.sheetnames == ["Sheet", "Sheet2"]
        assert book_xlsx["Sheet"]["A1"].value == 42
        assert book_xlsx["Sheet2"]["A1"].value == "test"
